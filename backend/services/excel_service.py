"""Server-side Excel export with styling, referencing docs/md_to_excel.py patterns.

Fixed: source_refs handling for both string and dict formats (2026-05-07)
"""
from __future__ import annotations

import io
import re
import traceback

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from backend.services import output_service
from backend.services.md_case_utils import format_numbered_multiline


COLUMNS = [
    ("#", 6),
    ("目录", 14),
    ("模块", 14),
    ("子项", 16),
    ("用例名称", 26),
    ("前置条件", 30),
    ("步骤", 40),
    ("预期结果", 30),
    ("优先级", 8),
    ("类型", 8),
    ("来源", 20),
    ("不确定", 8),
]

# 团队标准 10 列模板（与历史 Excel 对齐，便于团队直接合入回归库）
TEAM_TEMPLATE_COLUMNS = [
    ("用例目录", 14),
    ("模块", 14),
    ("子项", 18),
    ("用例名称", 26),
    ("前置条件", 30),
    ("用例步骤", 40),
    ("预期结果", 30),
    ("用例类型", 10),
    ("用例等级", 10),
    ("创建人", 12),
]

_TEAM_DEFAULT_COL_WIDTH = {name: w for name, w in TEAM_TEMPLATE_COLUMNS}

_DEFAULT_COL_WIDTH = {name: w for name, w in COLUMNS}


def _split_numbered(text: str) -> str:
    """将单元格内的编号内容分行，参考 docs/md_to_excel.py 的 split_numbered_content"""
    if not text:
        return text
    # 使用正则匹配 "空格+数字+点"，替换为 "换行+数字+点"
    return re.sub(r"\s+(\d+\.)", r"\n\1", str(text))


def _apply_style(ws, headers: list[str], n_data_rows: int,
                 width_map: dict | None = None,
                 priority_header: str = "优先级") -> None:
    """应用Excel样式，参考 docs/md_to_excel.py 的 apply_style"""
    n_cols = len(headers)
    total_rows = n_data_rows + 1

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    # 数据单元格样式
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # 交替行颜色
    even_fill = PatternFill("solid", fgColor="F5F0FF")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    
    # 优先级颜色
    priority_fills = {
        "高": PatternFill("solid", fgColor="FFE0E0"),
        "中": PatternFill("solid", fgColor="FFF5E0"),
        "低": PatternFill("solid", fgColor="E0F5E0"),
    }
    priority_col = next((i + 1 for i, h in enumerate(headers) if h == priority_header), None)

    # 应用表头样式
    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 应用数据样式
    for ri in range(2, total_rows + 1):
        row_fill = even_fill if ri % 2 == 0 else odd_fill
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = data_font
            cell.border = thin_border
            if ci == priority_col:
                pval = str(cell.value or "")
                if pval in priority_fills:
                    cell.fill = priority_fills[pval]
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.fill = row_fill
                    cell.alignment = data_align
            else:
                cell.fill = row_fill
                cell.alignment = data_align

    # 自动调整列宽（按换行后最长的一行计算）
    wmap = width_map if width_map is not None else _DEFAULT_COL_WIDTH
    for ci, col_name in enumerate(headers, 1):
        default_w = wmap.get(col_name, 14)
        max_len = len(col_name)
        for ri in range(2, total_rows + 1):
            val = str(ws.cell(row=ri, column=ci).value or "")
            if val:
                lines = val.split("\n")
                max_len = max(max_len, max(len(ln) for ln in lines))
        # 中文字符宽度调整，最大50
        adjusted_width = min(max_len * 2 + 2, 50)
        ws.column_dimensions[get_column_letter(ci)].width = max(adjusted_width, default_w)

    # 设置表头行高
    ws.row_dimensions[1].height = 25
    
    # 自动调整行高（根据单元格内换行数量）
    for ri in range(2, total_rows + 1):
        max_lines = 1
        for ci in range(1, n_cols + 1):
            val = str(ws.cell(row=ri, column=ci).value or "")
            if val:
                lines = val.count("\n") + 1
                max_lines = max(max_lines, lines)
        # 每行约18磅高度，最小20
        ws.row_dimensions[ri].height = max(20, max_lines * 18)

    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{total_rows}"


def export_testcase_excel(project: str, filename: str) -> bytes:
    """导出测试用例为Excel，支持JSON和Markdown格式"""
    import json
    
    try:
        raw = output_service.read_output_raw(project, "testcase", filename)
    except ValueError as e:
        raise ValueError(f"读取文件失败: {str(e)}")
    except Exception as e:
        error_msg = f"无法访问文件: {filename}, 错误: {str(e)}"
        print(f"[ERROR] {error_msg}\n{traceback.format_exc()}")
        raise ValueError(error_msg)

    # Markdown格式直接转换
    if filename.lower().endswith(".md"):
        try:
            return _excel_from_md(raw)
        except Exception as e:
            error_msg = f"Markdown转换失败: {str(e)}"
            print(f"[ERROR] {error_msg}\n{traceback.format_exc()}")
            raise ValueError(error_msg)

    # JSON格式解析
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        error_msg = f"JSON格式错误: {str(e)}, 文件: {filename}"
        print(f"[ERROR] {error_msg}")
        raise ValueError(error_msg)
    except Exception as e:
        error_msg = f"解析文件失败: {str(e)}"
        print(f"[ERROR] {error_msg}\n{traceback.format_exc()}")
        raise ValueError(error_msg)
    
    if not isinstance(data, dict):
        raise ValueError(f"数据格式错误: 期望JSON对象, 实际为 {type(data).__name__}")
    
    # 提取cases数组，支持多种嵌套结构
    cases = (data.get("cases") or 
             data.get("data", {}).get("cases") or 
             [])
    
    if not cases:
        raise ValueError("没有找到测试用例数据(cases数组为空)")
    
    try:
        return _excel_from_cases(cases)
    except Exception as e:
        error_msg = f"生成Excel失败: {str(e)}"
        print(f"[ERROR] {error_msg}\n{traceback.format_exc()}")
        raise ValueError(error_msg)


def _excel_from_md(md_content: str) -> bytes:
    from backend.services.md_case_utils import parse_md_table_rows, split_numbered, _SPLIT_COLS

    headers, rows = parse_md_table_rows(md_content)
    if not headers:
        raise ValueError("Markdown 中未找到有效表格")

    split_indices = {i for i, h in enumerate(headers) if h in _SPLIT_COLS}
    formatted_rows = []
    for cells in rows:
        new_cells = list(cells)
        for idx in split_indices:
            if idx < len(new_cells):
                new_cells[idx] = split_numbered(new_cells[idx])
        formatted_rows.append(new_cells)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(formatted_rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    _apply_style(ws, headers, len(formatted_rows))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def write_with_team_template(cases: list[dict], *, default_creator: str = "") -> bytes:
    """按团队标准 10 列模板导出。

    输入是 TestCase.model_dump() 形态（也容忍历史 LegacyCase 形态：title/case_type/case_id 等）。
    与 _excel_from_cases 区别：
      - 列与团队 Excel 对齐，便于直接合并入回归库
      - 不写 # 序号 / 来源 / 不确定 三列
      - 类型列默认填 "功能测试"（除非 cases 里给了 case_type/category）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    headers = [c[0] for c in TEAM_TEMPLATE_COLUMNS]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    for ri, c in enumerate(cases):
        # 步骤
        steps_arr = c.get("steps") or []
        if isinstance(steps_arr, list):
            steps_text = "\n".join(
                f"{(s.get('step') if isinstance(s, dict) else None) or (s.get('index') if isinstance(s, dict) else None) or j+1}."
                f" {(s.get('action', '') if isinstance(s, dict) else str(s))} "
                f"{(s.get('data', '') if isinstance(s, dict) else '')}".strip()
                for j, s in enumerate(steps_arr)
            )
        else:
            steps_text = str(steps_arr)

        # 前置
        pre_raw = c.get("preconditions", "") or c.get("preconditions_text", "")
        pre = format_numbered_multiline(pre_raw)

        # 预期：列表/单值/逐步合并
        expected_raw = c.get("expected_result") or c.get("expected") or ""
        expected = format_numbered_multiline(expected_raw)
        # 兜底：从 steps[*].expected 合并（LegacyCase 形态）
        if not expected and isinstance(steps_arr, list):
            expects = []
            for j, s in enumerate(steps_arr):
                if isinstance(s, dict) and s.get("expected"):
                    idx = s.get("index") or s.get("step") or j + 1
                    expects.append(f"{idx}. {s['expected']}")
            if expects:
                expected = "\n".join(expects)

        # 类型
        case_type = c.get("case_type") or c.get("type") or c.get("category") or "功能测试"
        # 等级（团队习惯：P0/P1/P2/P3 或 高/中/低）
        priority = c.get("priority", "") or ""

        # 创建人
        creator = c.get("creator") or default_creator or "auto"

        values = [
            c.get("suite") or c.get("catalog") or c.get("module", ""),  # 用例目录：缺则用 module
            c.get("module", ""),
            c.get("sub_item") or c.get("sub_item_base", ""),
            c.get("title") or c.get("name", ""),
            _split_numbered(pre),
            _split_numbered(steps_text),
            _split_numbered(expected),
            case_type,
            priority,
            creator,
        ]
        for ci, v in enumerate(values, 1):
            ws.cell(row=ri + 2, column=ci, value=v)

    _apply_style(
        ws, headers, len(cases),
        width_map=_TEAM_DEFAULT_COL_WIDTH,
        priority_header="用例等级",
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _excel_from_cases(cases: list[dict]) -> bytes:
    """从cases数组生成Excel，参考 docs/md_to_excel.py 的样式处理"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    headers = [c[0] for c in COLUMNS]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    for ri, c in enumerate(cases):
        # 处理steps：支持列表和字符串格式
        steps_arr = c.get("steps") or []
        if isinstance(steps_arr, list):
            steps_text = "\n".join(
                f"{s.get('step', j+1)}. {s.get('action', '')} {s.get('data', '')}".strip()
                if isinstance(s, dict)
                else f"{j+1}. {s}"
                for j, s in enumerate(steps_arr)
            )
        else:
            steps_text = str(steps_arr)

        # 处理preconditions：支持列表和字符串，标号分行
        preconditions = format_numbered_multiline(c.get("preconditions", ""))

        # 处理expected：兼容不同字段名，标号分行
        expected = format_numbered_multiline(c.get("expected") or c.get("expected_result", ""))

        # 处理source_refs：支持多种格式（字符串数组或字典数组）
        source_refs = c.get("source_refs") or []
        if isinstance(source_refs, list):
            sources_list = []
            for r in source_refs:
                if isinstance(r, dict):
                    # 字典格式：{"file": "...", "kp_id": "..."}
                    sources_list.append(r.get("file") or r.get("kp_id") or str(r))
                else:
                    # 字符串格式："文件名 #序号"
                    sources_list.append(str(r))
            sources = "; ".join(sources_list)
        else:
            sources = str(source_refs)

        values = [
            ri + 1,
            c.get("catalog", ""),
            c.get("module", ""),
            c.get("sub_item", ""),
            c.get("name") or c.get("title", ""),
            _split_numbered(preconditions),
            _split_numbered(steps_text),
            _split_numbered(expected),
            c.get("priority", ""),
            c.get("type") or c.get("category", ""),
            sources,
            "是" if c.get("uncertain") else "",
        ]
        for ci, v in enumerate(values, 1):
            ws.cell(row=ri + 2, column=ci, value=v)

    _apply_style(ws, headers, len(cases))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
