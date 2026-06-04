"""飞书集成业务编排层。

import_legacy: 从飞书多维表格/电子表格拉数据 → 物化成 xlsx 字节 →
              复用 legacy_service.ingest_excel 走老 pipeline。
              这样幂等、列映射确认、warning 收集全部沿用现有逻辑。

export_cases_to_sheet: 把项目内的 TestCase 列表渲染成固定列结构 →
                      调 client.create_sheet_with_records → 返回 share_url。
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, asdict
from typing import Any

from backend.integrations.feishu.client import (
    FeishuClient,
    SheetWriteResult,
    get_client,
)
from backend.integrations.feishu.config import load_config
from backend.schemas.column_mapping import ColumnMapping
from backend.services.legacy_service import ExcelIngestResult, ingest_excel

logger = logging.getLogger(__name__)


# F8 导出固定列顺序，避免下游消费方猜列。
SHEET_EXPORT_HEADERS = [
    "用例ID", "模块", "标题", "前置条件", "步骤",
    "预期结果", "优先级", "类别", "来源",
]


# ============ F1 导入 ============

def _records_to_xlsx_bytes(headers: list[str], records: list[dict[str, Any]]) -> bytes:
    """把 [{header: value}] 转成 xlsx 字节流。空缺单元格填空串。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for rec in records:
        ws.append([str(rec.get(h, "") or "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_legacy_from_feishu(
    project: str,
    url: str,
    *,
    confirmed_mapping: ColumnMapping | None = None,
    client: FeishuClient | None = None,
) -> ExcelIngestResult:
    """F1 主入口。

    幂等性由 legacy_service 保证（按字节 sha1）。同一份飞书表格不变 → 拉两次返回 already_parsed=True。
    """
    cli = client or get_client(project)
    pulled = cli.pull_bitable(url)
    if not pulled.records:
        logger.info("[feishu] import: empty records from %s", url)
    xlsx = _records_to_xlsx_bytes(pulled.headers, pulled.records)
    filename = f"feishu_{pulled.table_name}.xlsx"
    return ingest_excel(
        project=project,
        filename=filename,
        content=xlsx,
        confirmed_mapping=confirmed_mapping,
    )


# ============ F8 导出 ============

@dataclass
class CaseExportResult:
    sheet_token: str
    share_url: str
    row_count: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _case_to_row(case: dict[str, Any]) -> list[str]:
    """TestCase dict → SHEET_EXPORT_HEADERS 顺序的字符串行。"""
    steps = case.get("steps") or []
    steps_text = "\n".join(
        f"{i+1}. {s.get('action','')}" + (f"（数据：{s['data']}）" if s.get("data") else "")
        for i, s in enumerate(steps)
    )
    source_refs = case.get("source_refs") or []
    source_text = "; ".join(
        (r.get("file") or "")
        + (f"#{r['section']}" if r.get("section") else "")
        for r in source_refs
    )
    return [
        str(case.get("case_id", "")),
        str(case.get("feature_point", "")),
        str(case.get("title", "")),
        "\n".join(case.get("preconditions") or []),
        steps_text,
        str(case.get("expected_result", "")),
        str(case.get("priority", "")),
        str(case.get("category", "")),
        source_text,
    ]


def export_cases_to_sheet(
    project: str,
    cases: list[dict[str, Any]],
    *,
    title: str = "",
    client: FeishuClient | None = None,
) -> CaseExportResult:
    """F8 主入口。

    cases 期望为 TestCase.model_dump() 列表；若调用方传入 TestCase 对象，请先 dump。
    title 留空时按项目+时间戳生成。
    """
    cli = client or get_client(project)
    cfg = load_config(project)
    if not title:
        from backend.core.timeutil import utc_iso_z
        title = f"{project}-用例导出-{utc_iso_z()}"
    rows = [_case_to_row(c) for c in cases]
    result: SheetWriteResult = cli.create_sheet_with_records(
        title=title,
        headers=SHEET_EXPORT_HEADERS,
        rows=rows,
        folder_token=cfg.folder_token,
    )
    return CaseExportResult(
        sheet_token=result.sheet_token,
        share_url=result.share_url,
        row_count=result.row_count,
    )
