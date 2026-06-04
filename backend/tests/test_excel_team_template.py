"""excel_service.write_with_team_template 测试。

只验证：
  - 表头与团队 10 列模板一致
  - TestCase shape（steps/preconditions list、expected_result）正确写入
  - LegacyCase shape（steps[*].expected 合并为预期）兜底有效
  - 创建人列默认为 default_creator
"""
from __future__ import annotations

import io

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook

from backend.services import excel_service


EXPECTED_HEADERS = [
    "用例目录", "模块", "子项", "用例名称", "前置条件",
    "用例步骤", "预期结果", "用例类型", "用例等级", "创建人",
]


def _read_sheet(buf: bytes):
    wb = load_workbook(io.BytesIO(buf))
    ws = wb.active
    return ws


def test_headers_match_team_template():
    cases = [{"title": "x", "steps": [{"step": 1, "action": "a", "data": ""}],
              "expected_result": "ok", "module": "M", "priority": "P0"}]
    buf = excel_service.write_with_team_template(cases)
    ws = _read_sheet(buf)
    headers = [ws.cell(row=1, column=i + 1).value for i in range(len(EXPECTED_HEADERS))]
    assert headers == EXPECTED_HEADERS


def test_testcase_shape_writes_correctly():
    cases = [{
        "title": "登录-成功",
        "module": "账号",
        "sub_item": "登录-正常流程",
        "preconditions": ["已注册账号"],
        "steps": [
            {"step": 1, "action": "打开登录页", "data": ""},
            {"step": 2, "action": "输入账密", "data": "alice/123"},
        ],
        "expected_result": "跳转首页",
        "priority": "P0",
        "category": "正常",
    }]
    ws = _read_sheet(excel_service.write_with_team_template(cases, default_creator="bob"))

    row2 = [ws.cell(row=2, column=i + 1).value for i in range(len(EXPECTED_HEADERS))]
    assert row2[0] == "账号"           # 用例目录回退 module（cases 没给 suite/catalog）
    assert row2[1] == "账号"           # 模块
    assert row2[2] == "登录-正常流程"  # 子项
    assert row2[3] == "登录-成功"      # 用例名称
    assert "已注册账号" in row2[4]
    assert "1." in row2[5] and "2." in row2[5]
    assert "跳转首页" in row2[6]
    assert row2[7] == "正常"            # 类型回退到 category
    assert row2[8] == "P0"              # 等级
    assert row2[9] == "bob"             # 创建人


def test_legacy_shape_with_step_expecteds():
    """LegacyCase 形态：没有 expected_result，但 steps[*].expected 存在。"""
    cases = [{
        "title": "支付-下单",
        "module": "支付",
        "sub_item": "下单-正常流程",
        "preconditions": "已登录",
        "steps": [
            {"index": 1, "action": "选商品", "expected": "进入下单页"},
            {"index": 2, "action": "确认订单", "expected": "提示成功"},
        ],
        "case_type": "功能测试",
        "priority": "P1",
        "creator": "alice",
    }]
    ws = _read_sheet(excel_service.write_with_team_template(cases))
    row2 = [ws.cell(row=2, column=i + 1).value for i in range(len(EXPECTED_HEADERS))]
    assert row2[7] == "功能测试"
    assert row2[9] == "alice"
    expected_cell = row2[6]
    assert "进入下单页" in expected_cell
    assert "提示成功" in expected_cell


def test_default_type_when_missing():
    cases = [{"title": "x", "steps": [{"step": 1, "action": "a"}],
              "expected_result": "ok"}]
    ws = _read_sheet(excel_service.write_with_team_template(cases))
    type_cell = ws.cell(row=2, column=8).value
    assert type_cell == "功能测试"


def test_empty_cases_writes_only_header():
    ws = _read_sheet(excel_service.write_with_team_template([]))
    assert ws.cell(row=1, column=1).value == "用例目录"
    assert ws.cell(row=2, column=1).value is None
