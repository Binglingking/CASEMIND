import pandas as pd
import markdown
import re
from bs4 import BeautifulSoup
from io import StringIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def split_numbered_content(text):
    """
    将单元格内的编号内容分行（等同于VBA的 单元格内分行 功能）
    例如："1.登录 2.确认授权 3.查看数据" → "1.登录\n2.确认授权\n3.查看数据"

    :param text: 原始文本
    :return: 处理后的文本（带换行符）
    """
    if pd.isna(text) or text == '':
        return text

    text = str(text)
    # 使用正则匹配 "空格+数字+点"，替换为 "换行+数字+点"
    # 支持任意位数编号（1. 2. ... 10. 11. ...）
    text = re.sub(r'\s+(\d+\.)', r'\n\1', text)

    return text


def md_to_excel_pretty(md_file, excel_file):
    """带样式美化的MD转Excel（含单元格内自动分行）"""
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()

    html = markdown.markdown(content, extensions=['tables'])
    soup = BeautifulSoup(html, 'lxml')
    tables = soup.find_all('table')

    if not tables:
        print("❌ 未找到表格")
        return

    # ✨ 定义需要分行处理的目标列（支持模糊匹配列名）
    target_columns = ['前置条件', '用例步骤', '预期结果']

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for i, table in enumerate(tables):
            sheet_name = f'Sheet{i + 1}'
            prev = table.find_previous(['h1', 'h2', 'h3'])
            if prev:
                sheet_name = prev.get_text().strip()[:31]
                sheet_name = ''.join(c for c in sheet_name if c not in '[]:*?/\\')

            df = pd.read_html(StringIO(str(table)))[0]

            # ✨ 只对指定列进行分行处理（模糊匹配列名）
            processed_cols = []
            for col in df.columns:
                if any(keyword in str(col) for keyword in target_columns):
                    df[col] = df[col].apply(split_numbered_content)
                    processed_cols.append(col)

            if processed_cols:
                print(f"📝 [{sheet_name}] 已对列进行分行处理: {processed_cols}")

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取worksheet应用样式
            ws = writer.sheets[sheet_name]
            apply_style(ws, df)

    print(f"🎉 转换完成（已美化+自动分行）：{excel_file}")


def apply_style(ws, df):
    """应用Excel样式"""
    # 表头样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center', vertical='center')

    # 边框样式
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # 数据单元格样式
    data_font = Font(name='微软雅黑', size=10)
    # ✨ vertical改为top，换行内容从顶部开始显示更美观
    data_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # 应用表头样式
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 应用数据样式
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # 自动调整列宽（按换行后最长的一行计算，避免列过宽）
    for col_idx, col_name in enumerate(df.columns, 1):
        col_data = df.iloc[:, col_idx - 1].astype(str)
        max_length = len(str(col_name))

        for val in col_data:
            if val and val != 'nan':
                # 按换行符拆分后，取最长的一行
                lines = str(val).split('\n')
                line_max = max(len(line) for line in lines)
                max_length = max(max_length, line_max)

        # 中文字符宽度调整
        adjusted_width = min(max_length * 2 + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # ✨ 自动调整行高（根据单元格内换行数量）
    for row_idx in range(2, len(df) + 2):
        max_lines = 1
        for col_idx in range(1, len(df.columns) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                lines = str(cell_value).count('\n') + 1
                max_lines = max(max_lines, lines)
        # 每行约18磅高度，最小20
        ws.row_dimensions[row_idx].height = max(20, max_lines * 18)

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置表头行高
    ws.row_dimensions[1].height = 25


# 使用示例
if __name__ == '__main__':
    md_to_excel_pretty(
        r'/docs/AI用例/230021素材归因及报表_测试用例.md',
        r'D:\CaseMind\docs\AI用例\230021素材归因及报表_测试用例pro.xlsx'
    )